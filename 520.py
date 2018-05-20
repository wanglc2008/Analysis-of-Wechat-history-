#!/usr/bin/python
# -*- coding: UTF-8 -*-
from openpyxl import load_workbook
import sys
import io
import matplotlib.pyplot as plt
from wordcloud import WordCloud, STOPWORDS, ImageColorGenerator
import numpy as np
from PIL import Image

def excel_and_write():
    wb = load_workbook("message.xlsx")
    print(wb.sheetnames)
    sheet = wb.get_sheet_by_name('Sheet1')
    max_row=sheet.max_row
    f=open('test.txt','a+',encoding='utf-8')
    for num in range(2,max_row+1):
        num=str(num)
        str_path='A'+num
        str_message=str(sheet[str_path].value)
        if len(str_message)<23 :
            f.write(str_message)
            print(sheet[str_path].value)
    f.close()

def getpic(path):
    f=open('test.txt','r+',encoding='utf-8')
    path=str(path)
    alice_coloring = np.array(Image.open(path+'.png'))
    text=f.read()
    stopwords = set(STOPWORDS)
    stopwords.add("微笑")
    stopwords.add("流泪")
    font = r'C:\Windows\Fonts\simhei.ttf'
    wc = WordCloud( font_path=font,background_color="white", stopwords=stopwords,max_words=200000000000,max_font_size=70,mask=alice_coloring).generate(text.lower())
    image_colors = ImageColorGenerator(alice_coloring)
    plt.imshow(wc)
    plt.axis("off")
    plt.show()
    wc.to_file(path+'_x.png')

if __name__ == '__main__':

    excel_and_write()
    for x in [1,2,3]:
        getpic(x)
